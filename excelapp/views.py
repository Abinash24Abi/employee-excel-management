from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import UploadExcelForm
from .models import Employee
import openpyxl

from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import units
import pandas as pd
import pypandoc
from io import BytesIO


def home(request):
    employees = Employee.objects.all()
    return render(request, "home.html", {"employees": employees})

# def upload_excel(request):
#     if request.method == "POST":
#         form = UploadExcelForm(request.POST, request.FILES)

#         if form.is_valid():
#             excel_file = request.FILES['excel_file']

#             # Validate file type
#             if not excel_file.name.endswith('.xlsx'):
#                 messages.error(request, "Please upload a valid Excel (.xlsx) file.")
#                 return redirect('upload_excel')

#             wb = openpyxl.load_workbook(excel_file)
#             sheet = wb.active

#             inserted = 0
#             skipped = 0

#             for row in sheet.iter_rows(min_row=2, values_only=True):
#                 emp_id, name, email, department = row

#                 if not Employee.objects.filter(emp_id=emp_id).exists():
#                     Employee.objects.create(
#                         emp_id=emp_id,
#                         name=name,
#                         email=email,
#                         department=department
#                     )
#                     inserted += 1
#                 else:
#                     skipped += 1

#             messages.success(
#                 request,
#                 f"Upload completed! Inserted: {inserted}, Skipped (duplicates): {skipped}"
#             )
#             return redirect('upload_excel')
#         else:
#             messages.error(request, "Invalid form submission.")

#     else:
#         form = UploadExcelForm()

#     return render(request, "upload.html", {"form": form})




def upload_excel(request):
    if request.method == "POST":
        form = UploadExcelForm(request.POST, request.FILES)

        if form.is_valid():
            excel_file = request.FILES['excel_file']

            if not excel_file.name.endswith('.xlsx'):
                messages.error(request, "Please upload a valid Excel (.xlsx) file.")
                return redirect('upload_excel')

            try:
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active

                inserted = 0
                updated = 0
                skipped = 0

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    emp_id, name, email, department = row

                    if not emp_id:
                        continue

                    try:
                        employee = Employee.objects.get(emp_id=emp_id)

                        if (
                            employee.name != name or
                            employee.email != email or
                            employee.department != department
                        ):
                            employee.name = name
                            employee.email = email
                            employee.department = department
                            employee.save()
                            updated += 1
                        else:
                            skipped += 1

                    except Employee.DoesNotExist:
                        Employee.objects.create(
                            emp_id=emp_id,
                            name=name,
                            email=email,
                            department=department
                        )
                        inserted += 1

                total = inserted + updated + skipped

                # Store summary in session
                request.session['upload_summary'] = {
                    'inserted': inserted,
                    'updated': updated,
                    'skipped': skipped,
                    'total': total
                }

                return redirect('upload_summary')

            except Exception as e:
                messages.error(request, f"Error processing file: {str(e)}")
                return redirect('upload_excel')

        else:
            messages.error(request, "Invalid form submission.")
            return redirect('upload_excel')

    else:
        form = UploadExcelForm()

    return render(request, "upload.html", {"form": form})


def upload_summary(request):
    summary = request.session.get('upload_summary')

    if not summary:
        return redirect('home')

    return render(request, "upload_summary.html", summary)



def export_pdf(request):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer)
    elements = []

    employees = Employee.objects.all()

    data = [["Emp ID", "Name", "Email", "Department"]]

    for emp in employees:
        data.append([emp.emp_id, emp.name, emp.email, emp.department])

    table = Table(data)
    table.setStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])

    elements.append(table)
    doc.build(elements)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="employees.pdf"'

    return response



def export_excel(request):
    employees = Employee.objects.all().values()

    df = pd.DataFrame(list(employees))

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employees.xlsx'

    df.to_excel(response, index=False)

    return response




def export_csv(request):
    employees = Employee.objects.all().values()

    df = pd.DataFrame(list(employees))

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=employees.csv'

    df.to_csv(response, index=False)

    return response



def export_txt(request):
    employees = Employee.objects.all()

    text_data = "Employee List\n\n"

    for emp in employees:
        text_data += f"{emp.emp_id} | {emp.name} | {emp.email} | {emp.department}\n"

    response = HttpResponse(content_type='text/plain')
    response['Content-Disposition'] = 'attachment; filename=employees.txt'
    response.write(text_data)

    return response